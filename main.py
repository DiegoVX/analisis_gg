# main.py
import tkinter as tk
import re
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from tkinter import ttk, messagebox, filedialog
from sql_checker import buscar_coincidencia_siadal
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import ttk as ttk_native
import sys
import os

# Interfaz gráfica
ventana = tk.Tk()
ventana.title("Validación de Materiales con SIADAL")
ventana.geometry("1100x700")
ventana.configure(bg="#f8f9fa")

# Título principal
titulo = tk.Label(
    ventana,
    text="Validación de Materiales con SIADAL",
    font=("Segoe UI", 18, "bold"),
    bg="#f8f9fa",
    fg="#0d6efd"
)
titulo.pack(fill="x", padx=20, pady=(18, 5))

# Variables globales
df_detalle = None
archivo_actual = tk.StringVar(value="No se ha cargado ningún archivo.")

# Label para mostrar archivo cargado
frame_archivo = tk.Frame(ventana, bg="#e9ecef", bd=1, relief="ridge")
frame_archivo.pack(fill="x", padx=30, pady=(0, 10))
label_archivo = tk.Label(
    frame_archivo,
    textvariable=archivo_actual,
    font=("Segoe UI", 11, "italic"),
    bg="#e9ecef",
    fg="#0d6efd",
    anchor="w",
    padx=10,
    pady=6
)
label_archivo.pack(fill="x")

# Sidebar para los botones (estilo Bootstrap 5)
sidebar = tk.Frame(ventana, width=220, bg="#f8f9fa")
sidebar.pack(side="left", fill="y", padx=(10, 0), pady=10)

estilo_btn = {
    'font': ("Segoe UI", 11, "bold"),
    'bd': 0,
    'relief': 'flat',
    'activebackground': '#e2e6ea',
    'cursor': 'hand2',
    'height': 2,
    'width': 20,
    'highlightthickness': 0,
}

# --- Widgets y tabla ---
frame_tabla = tk.Frame(ventana, bg="#f8f9fa")
frame_tabla.pack(fill="both", expand=True, padx=(0, 30), pady=20)

scroll_y = tk.Scrollbar(frame_tabla, orient="vertical")
scroll_x = tk.Scrollbar(frame_tabla, orient="horizontal")

tabla = ttk.Treeview(
    frame_tabla,
    columns=["n", "material", "factura", "cantidad", "material_siadal", "estado"],
    show="headings",
    height=22,
    yscrollcommand=scroll_y.set,
    xscrollcommand=scroll_x.set
)

headers = [
    ("n", "N°"),
    ("material", "Número Material"),
    ("factura", "Número Factura"),
    ("cantidad", "Cantidad UMC"),
    ("material_siadal", "Material SIADAL"),
    ("estado", "Estado")
]

for col, txt in headers:
    tabla.heading(col, text=txt)
    tabla.column(col, anchor="center", width=160 if col != "n" else 60)

scroll_y.config(command=tabla.yview)
scroll_x.config(command=tabla.xview)
scroll_y.pack(side="right", fill="y")
scroll_x.pack(side="bottom", fill="x")
tabla.pack(fill="both", expand=True)

# Barra de progreso global (debe ir después de crear ventana)
progress_var = tk.DoubleVar()
progress_bar = ttk_native.Progressbar(ventana, variable=progress_var, maximum=100, length=400)
progress_label = tk.Label(ventana, text="")

# --- Funciones de evento y lógica ---
# --- Botón y función para exportar el análisis a Excel ---
def exportar_excel():
    import openpyxl
    from openpyxl.styles import PatternFill

    # Exporta los datos que se muestran en la tabla, no el DataFrame original
    columnas_tabla = [tabla.heading(col)["text"] for col in tabla["columns"]]
    datos_tabla = []
    for item in tabla.get_children():
        fila = tabla.item(item)["values"]
        datos_tabla.append(fila)
    if not datos_tabla:
        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
        return
    df_export = pd.DataFrame(datos_tabla, columns=columnas_tabla)
    archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not archivo:
        return
    try:
        df_export.to_excel(archivo, index=False, engine="openpyxl")
        # --- Pintar celdas vacías en rojo ---
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        # Buscar la columna "Material SIADAL"
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Material SIADAL":
                col_idx = idx
                break
        if col_idx:
            rojo = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.value = "REVISAR DE FORMA MANUAL"
                        cell.fill = rojo
        wb.save(archivo)
        messagebox.showinfo("Éxito", f"Archivo exportado correctamente:\n{archivo}")
    except PermissionError:
        messagebox.showerror("Error", "El archivo está abierto o protegido. Ciérralo e inténtalo de nuevo.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar el archivo:\n{e}")

# --- Función para actualizar el excel y generar bitacora ---
def actualizar_parciales_y_bitacora():
    archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not archivo:
        return

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    # Buscar columnas
    col_material_siadal = None
    col_estado = None
    col_n = None
    col_numero_material = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Material SIADAL":
            col_material_siadal = idx
        if cell.value == "Estado":
            col_estado = idx
        if cell.value == "N°":
            col_n = idx
        if cell.value == "Número Material":
            col_numero_material = idx

    if not col_material_siadal or not col_estado or not col_numero_material:
        messagebox.showerror("Error", "No se encontraron las columnas necesarias en el archivo.")
        return

    bitacora = []
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Recorrer filas y actualizar donde el estado sea "Coincidencia Parcial"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        estado_cell = row[col_estado - 1]
        siadal_cell = row[col_material_siadal - 1]
        n_cell = row[col_n - 1] if col_n else None
        numero_material_cell = row[col_numero_material - 1]
        if str(estado_cell.value).strip() == "Coincidencia Parcial":
            fila_excel = estado_cell.row
            # Buscar en la tabla visual por número de fila (N°)
            for item in tabla.get_children():
                fila_tabla = tabla.item(item)["values"]
                if int(fila_tabla[0]) == fila_excel - 1 and fila_tabla[5] == "Coincidencia Parcial":
                    valor_anterior = numero_material_cell.value  # Valor anterior del Excel
                    valor_nuevo = fila_tabla[4]  # Valor de SIADAL
                    siadal_cell.value = valor_nuevo
                    siadal_cell.fill = amarillo
                    numero_material_cell.value = valor_nuevo
                    celda_excel = f"{get_column_letter(col_material_siadal)}{fila_excel}"
                    celda_numero_material = f"{get_column_letter(col_numero_material)}{fila_excel}"
                    bitacora.append({
                        "Celda Actualizada (Material SIADAL)": celda_excel,
                        "Celda Actualizada (Número Material)": celda_numero_material,
                        "Valor Anterior": valor_anterior,
                        "Valor Nuevo": valor_nuevo,
                        "Número Factura": fila_tabla[2]
                    })
                    break

    wb.save(archivo)
    messagebox.showinfo("Éxito", "Excel actualizado correctamente.")

    # Guardar la bitácora
    if bitacora:
        df_bitacora = pd.DataFrame(bitacora)
        archivo_bitacora = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar bitácora de actualización"
        )
        if archivo_bitacora:
            df_bitacora.to_excel(archivo_bitacora, index=False)
            messagebox.showinfo("Bitácora", f"Bitácora guardada en:\n{archivo_bitacora}")
    else:
        messagebox.showinfo("Bitácora", "No hubo coincidencias parciales para actualizar.")

# --- Función para filtrar importaciones ---
def filtrar_importaciones(ruta_archivo):
    """
    Carga la hoja 'ENCABEZADO FAC' y filtra solo los registros de importación (Tipo Operación == 1).
    Retorna el DataFrame filtrado, sin modificar el original.
    Si la hoja no existe, retorna un DataFrame vacío.
    """
    df = pd.read_excel(ruta_archivo, sheet_name="ENCABEZADO FAC")
    df_import = df[df["Tipo Operación (1. Importación 2.Exportación)"] == 1]
    return df_import

# --- Botón y función para mostrar estadísticos ---
def mostrar_estadisticos():
    estados = [tabla.item(item)["values"][5] for item in tabla.get_children()]
    total = len(estados)
    exacta = estados.count("Coincidencia Exacta")
    parcial = estados.count("Coincidencia Parcial")
    sin = estados.count("Sin Coincidencia")
    win = tk.Toplevel(ventana)
    win.title("Estadísticos de Coincidencias")
    win.geometry("350x220")
    win.configure(bg="#f8f9fa")
    tk.Label(win, text=f"Total de materiales: {total}", font=("Segoe UI", 12, "bold"), bg="#f8f9fa", fg="#212529").pack(
        pady=(20, 10))
    tk.Label(win, text=f"Coincidencia Exacta: {exacta}", font=("Segoe UI", 12), bg="#f8f9fa", fg="#198754").pack(pady=5)
    tk.Label(win, text=f"Coincidencia Parcial: {parcial}", font=("Segoe UI", 12), bg="#f8f9fa", fg="#ffc107").pack(
        pady=5)
    tk.Label(win, text=f"Sin Coincidencia: {sin}", font=("Segoe UI", 12), bg="#f8f9fa", fg="#dc3545").pack(pady=5)
    tk.Button(win, text="Cerrar", command=win.destroy, bg="#0d6efd", fg="white", font=("Segoe UI", 11, "bold"), bd=0,
              relief='flat', activebackground="#0b5ed7", cursor='hand2', height=1, width=12).pack(pady=15)

# --- Función para mostrar gráfico de pastel ---
def mostrar_grafico_pastel():
    estados = [tabla.item(item)["values"][5] for item in tabla.get_children()]
    labels = ["Coincidencia Exacta", "Coincidencia Parcial", "Sin Coincidencia"]
    sizes = [estados.count(l) for l in labels]
    colors = ["#198754", "#ffc107", "#dc3545"]
    if sum(sizes) == 0:
        messagebox.showinfo("Sin datos", "No hay datos para graficar.")
        return
    win = tk.Toplevel(ventana)
    win.title("Gráfico de Coincidencias")
    win.geometry("420x420")
    fig, ax = plt.subplots(figsize=(4, 4))
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors, startangle=90,
                                      textprops={'fontsize': 11, 'fontname': 'Segoe UI'})
    ax.axis("equal")
    plt.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=win)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)
    tk.Button(win, text="Cerrar", command=win.destroy, bg="#0d6efd", fg="white", font=("Segoe UI", 11, "bold"), bd=0,
              relief='flat', activebackground="#0b5ed7", cursor='hand2', height=1, width=12).pack(pady=10)

# Función de normalización avanzada para comparar materiales
def normalizar_material(mat):
    if not isinstance(mat, str):
        mat = str(mat)
    mat = mat.strip().upper()
    mat = re.sub(r'[\s\-.\/]', '', mat)  # Quita espacios, guiones, puntos, diagonales
    mat = mat.lstrip('0')
    return mat

# Mostrar datos en tabla con colores y coincidencias
def mostrar_materiales(df):
    for row in tabla.get_children():
        tabla.delete(row)

    for idx, row in df.iterrows():
        numero_material = str(row["Número Material"]).strip()
        numero_factura = str(row["Número Factura"]).strip()
        cantidad_umc = row["Cantidad UMC"]

        resultado = buscar_coincidencia_siadal(numero_material, numero_factura, cantidad_umc)
        match = resultado["match"]
        material_siadal = resultado.get("matno", "")

        if match == "exacto":
            color_tag = "verde"
            estado = "Coincidencia Exacta"
        elif match == "parcial":
            color_tag = "amarillo"
            estado = "Coincidencia Parcial"
        else:
            color_tag = "rojo"
            estado = "Sin Coincidencia"
            if not material_siadal:
                np_siadal = row.get("NP SIADAL", "")
                # Si NP SIADAL es válido, usarlo
                if not pd.isna(np_siadal) and str(np_siadal).strip().lower() != "nan":
                    material_siadal = str(np_siadal).strip()
                else:
                    # 1. Buscar coincidencia por número de parte y cantidad (no factura)
                    filtro_np_cant = df[
                        (df["Número Material"].astype(str).str.strip() == numero_material) &
                        (df["Cantidad UMC"] == cantidad_umc) &
                        (df["Número Factura"].astype(str).str.strip() != numero_factura)
                        ]
                    if not filtro_np_cant.empty:
                        fila = filtro_np_cant.iloc[0]
                        almacen = fila.get("Almacen", "")
                        if almacen and not pd.isna(almacen):
                            material_siadal = f"{fila['Número Material']} (Almacén: {almacen})"
                        else:
                            material_siadal = str(fila['Número Material'])
                    else:
                        # 2. Buscar coincidencia por factura y cantidad (no número de parte)
                        filtro_fac_cant = df[
                            (df["Número Factura"].astype(str).str.strip() == numero_factura) &
                            (df["Cantidad UMC"] == cantidad_umc) &
                            (df["Número Material"].astype(str).str.strip() != numero_material)
                            ]
                        if not filtro_fac_cant.empty:
                            fila = filtro_fac_cant.iloc[0]
                            almacen = fila.get("Almacen", "")
                            if almacen and not pd.isna(almacen):
                                material_siadal = f"{fila['Número Material']} (Almacén: {almacen})"
                            else:
                                material_siadal = str(fila['Número Material'])
                        else:
                            # 3. Buscar coincidencia solo por factura (ya lo tienes)
                            filtro = df[df["Número Factura"].astype(str).str.strip() == numero_factura]
                            if not filtro.empty:
                                for _, fila in filtro.iterrows():
                                    mat_encontrado = str(fila["Número Material"]).strip()
                                    if mat_encontrado != numero_material:
                                        almacen = fila.get("Almacen", "")
                                        if almacen and not pd.isna(almacen):
                                            material_siadal = f"{mat_encontrado} (Almacén: {almacen})"
                                        else:
                                            material_siadal = mat_encontrado
                                        break
                            else:
                                material_siadal = ""

        tabla.insert("", "end",
                     values=(idx + 1, numero_material, numero_factura, cantidad_umc, material_siadal, estado),
                     tags=(color_tag,))

    tabla.tag_configure("verde", background="lightgreen")
    tabla.tag_configure("amarillo", background="khaki")
    tabla.tag_configure("rojo", background="salmon")


# Evento para cargar Excel con barra de progreso
def evento_cargar():
    global df_detalle
    archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not archivo:
        return
    # Actualizar el label con el nombre del archivo
    archivo_actual.set(f"Archivo cargado: {archivo.split('/')[-1] if '/' in archivo else archivo.split('\\\\')[-1]}")
    progress_var.set(0)
    progress_bar.pack(pady=(20, 0))
    progress_label.pack()
    ventana.update()
    try:
        # Intentar leer la hoja 'DETALLE FAC', si no existe, leer la primera hoja disponible
        try:
            df_tmp = pd.read_excel(archivo, sheet_name="DETALLE FAC", engine="openpyxl")
        except Exception:
            xl = pd.ExcelFile(archivo, engine="openpyxl")
            primera_hoja = xl.sheet_names[0]
            df_tmp = xl.parse(primera_hoja)
        # --- FILTRAR SOLO IMPORTACIONES ---
        try:
            df_import = filtrar_importaciones(archivo)
            facturas_import = df_import["Número Factura"].astype(str).unique()
            df_tmp = df_tmp[df_tmp["Número Factura"].astype(str).isin(facturas_import)]
        except Exception as e:
            messagebox.showwarning("Advertencia", f"No se pudo filtrar importaciones:\n{e}")
        total = len(df_tmp)
        if total == 0:
            raise Exception("El archivo no contiene datos.")
        df_detalle = pd.DataFrame()
        for i, row in df_tmp.iterrows():
            df_detalle = pd.concat([df_detalle, pd.DataFrame([row])], ignore_index=True)
            percent = (i + 1) / total * 100
            progress_var.set(percent)
            progress_label.config(text=f"Cargando: {percent:.1f}%")
            ventana.update_idletasks()  # Refresca la barra de progreso
        # Asegura que la barra llegue a 100%
        progress_var.set(100)
        progress_label.config(text="Carga completada.")
        ventana.update_idletasks()
        mostrar_materiales(df_detalle)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el archivo:\n{e}")
        progress_label.config(text="Error al cargar.")
    finally:
        ventana.after(1200, lambda: (progress_bar.pack_forget(), progress_label.pack_forget()))


def reiniciar_programa():
    """Reinicia el programa actual."""
    python = sys.executable
    os.execl(python, python, *sys.argv)


def actualizar_ventana():
    """Limpia la tabla y reinicia los datos mostrados en la ventana."""
    global df_detalle
    for row in tabla.get_children():
        tabla.delete(row)
    df_detalle = None
    archivo_actual.set("No se ha cargado ningún archivo.")
    progress_label.config(text="Ventana actualizada. Lista para cargar nuevos datos.")


# --- Botones (después de todo lo anterior) ---
btn_cargar = tk.Button(sidebar, text="Cargar Excel", command=evento_cargar, bg="#0d6efd", fg="white",
                       activeforeground="white", **estilo_btn)
btn_cargar.pack(pady=(30, 12), padx=16)

btn_exportar = tk.Button(sidebar, text="Exportar Excel", command=exportar_excel, bg="#198754", fg="white",
                         activeforeground="white", **estilo_btn)
btn_exportar.pack(pady=12, padx=16)

btn_actualizar_parcial = tk.Button(
    sidebar,
    text="Actualizar Excel",
    command=actualizar_parciales_y_bitacora,
    bg="#fd7e14",
    fg="white",
    activeforeground="white",
    **estilo_btn
)
btn_actualizar_parcial.pack(pady=12, padx=16)

btn_actualizar = tk.Button(
    sidebar,
    text="Actualizar Programa",
    command=actualizar_ventana,
    bg="#0dcaf0",
    fg="white",
    activeforeground="white",
    **estilo_btn
)
btn_actualizar.pack(pady=12, padx=16)

btn_estadisticos = tk.Button(sidebar, text="Ver Estadísticos", command=mostrar_estadisticos, bg="#ffc107",
                             fg="#212529", activeforeground="#212529", **estilo_btn)
btn_estadisticos.pack(pady=12, padx=16)

btn_grafico = tk.Button(sidebar, text="Gráfico de Pastel", command=lambda: mostrar_grafico_pastel(), bg="#dc3545",
                        fg="white", activeforeground="white", **estilo_btn)
btn_grafico.pack(pady=12, padx=16)

btn_salir = tk.Button(sidebar, text="Salir", command=ventana.destroy, bg="#6c757d",
                      fg="white", activeforeground="white", **estilo_btn)
btn_salir.pack(pady=(40, 12), padx=16)

ventana.mainloop()