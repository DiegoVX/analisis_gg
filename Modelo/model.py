import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class DataModel:
    """Modelo para manejar datos de Excel y consultas SQL."""

    def __init__(self):
        self.df_original = None
        self.materiales_siadal = set()
        self.materiales_encontrados_avanzados = set()
        self.materiales_vistos = set()

    def cargar_excel(self, archivo):
        if not archivo:
            return False, "No se seleccionó ningún archivo."

        try:
            detalle_df = pd.read_excel(archivo, sheet_name="DETALLE FAC", engine="openpyxl")
            relacion_df = pd.read_excel(archivo, sheet_name="RELACIÓN FAC-PED", engine="openpyxl", header=None)
            relacion_df.columns = relacion_df.iloc[0]
            relacion_df = relacion_df[1:][["NumeroFactura", "NumeroPedimento"]].drop_duplicates(
                subset=["NumeroFactura"])

            encabezado_df = pd.read_excel(archivo, sheet_name="ENCABEZADO FAC", engine="openpyxl", header=None)
            encabezado_df.columns = ["NumeroFactura", "?", "TipoOperacion"] + list(encabezado_df.columns[3:])
            encabezado_df = encabezado_df[1:][["NumeroFactura", "TipoOperacion"]].drop_duplicates(
                subset=["NumeroFactura"])

            detalle_df["Número Factura"] = detalle_df["Número Factura"].astype(str)
            relacion_df["NumeroFactura"] = relacion_df["NumeroFactura"].astype(str)
            encabezado_df["NumeroFactura"] = encabezado_df["NumeroFactura"].astype(str)

            merged_df = detalle_df.merge(relacion_df, left_on="Número Factura", right_on="NumeroFactura", how="left")
            merged_df = merged_df.merge(encabezado_df, left_on="Número Factura", right_on="NumeroFactura", how="left")
            merged_df["TipoOperacion"] = merged_df["TipoOperacion"].map({1: "Importación", 2: "Exportación"})

            self.df_original = merged_df[
                ["Número Material", "Número Factura", "Cantidad UMC", "NumeroPedimento", "TipoOperacion"]]

            return True, "Archivo cargado correctamente."
        except Exception as e:
            return False, f"No se pudo procesar el archivo: {e}"

    def filtrar_datos(self, tipo_operacion):
        if self.df_original is None:
            return None
        if tipo_operacion == "Todos":
            return self.df_original
        return self.df_original[self.df_original["TipoOperacion"] == tipo_operacion]

    def consultar_sql(self):
        if self.df_original is None:
            return False, "No hay datos cargados."

        try:
            conexion = pyodbc.connect(
                'DRIVER={ODBC Driver 17 for SQL Server};'
                'SERVER=PRACTICAS_TI\\MSSQLSERVER1;'
                'DATABASE=dbSiadalGoGlobal;'
                'UID=sa;PWD=root'
            )
            cursor = conexion.cursor()

            cursor.execute("SELECT MatNoParte FROM siadalgoglobaluser.tblMaterial")
            self.materiales_siadal = set(str(r[0]).strip() for r in cursor.fetchall() if r[0])
            self.materiales_vistos = set(str(x).strip() for x in self.df_original["Número Material"].dropna().unique())

            return True, "Consulta SQL completada."
        except Exception as e:
            return False, f"Error: {e}"

    def buscar_coincidencias_avanzadas(self, tipo_operacion):
        if self.df_original is None or not self.materiales_siadal:
            return False, "Datos no cargados o consulta SQL no realizada.", []

        df_filtrado = self.filtrar_datos(tipo_operacion)

        df_no_encontrados = df_filtrado[
            ~df_filtrado["Número Material"].astype(str).str.strip().isin(self.materiales_siadal)
        ]

        if df_no_encontrados.empty:
            return True, "No hay materiales no encontrados.", []

        try:
            conexion = pyodbc.connect(
                'DRIVER={ODBC Driver 17 for SQL Server};'
                'SERVER=PRACTICAS_TI\\MSSQLSERVER1;'
                'DATABASE=dbSiadalGoGlobal;'
                'UID=sa;PWD=root'
            )
            cursor = conexion.cursor()

            resultados_totales = []
            duplicados_verificados = set()

            for fila in df_no_encontrados.itertuples():
                mat = str(fila._1).strip().upper().replace("'", "''")
                factura = str(fila._2).strip().replace("'", "''")
                try:
                    cantidad_excel = float(fila._3)
                except (ValueError, TypeError):
                    continue

                # NUEVO: tolerancia en comparación de cantidades (±5%)
                tolerancia = 0.05

                query = f"""
                SELECT
                    c.FactEntFechaEnt AS FECHA, p.ProvNombre, c.FactEntPedimento, c.FactEntNofact, 
                    c.FactEntFolio, c.FactEntContenedor AS Caja_CTR, a.MatNoParte, a.MatDescr, 
                    SUM(b.MovExistente) AS qty
                FROM siadalgoglobaluser.tblMaterial AS a
                INNER JOIN siadalgoglobaluser.tblMovimientos AS b ON a.idMaterial = b.idMaterial
                INNER JOIN siadalgoglobaluser.tblFacturaEnt AS c ON b.idFactEnt = c.idFactEnt
                INNER JOIN siadalgoglobaluser.tblDescrFactEnt AS d ON a.idMaterial = d.idMaterial 
                    AND c.idFactEnt = d.idFactEnt AND b.iddescrFERenTras = d.idDescrFactEnt
                INNER JOIN siadalgoglobaluser.tblProveedor AS p ON c.idProveedor = p.idProveedor
                WHERE c.idAlmacen IN (17, 8, 15)
                  AND c.FactEntFechaEnt >= 20240101
                  AND c.FactEntNofact = '{factura}'
                  AND a.MatNoParte LIKE '{mat}%'
                GROUP BY
                    c.FactEntFechaEnt, p.ProvNombre, c.FactEntFechaFactura, c.FactEntPedimento, 
                    c.FactEntNofact, c.FactEntFolio, c.FactEntContenedor, a.MatDescr, a.MatNoParte
                ORDER BY FECHA DESC
                """

                cursor.execute(query)
                for r in cursor.fetchall():
                    mat_encontrado = str(r.MatNoParte).strip().upper()
                    clave = (mat_encontrado, r.FactEntNofact)

                    # Validación por cantidad con tolerancia del 5%
                    cantidad_siadal = float(r.qty)
                    if abs(cantidad_siadal - cantidad_excel) / max(cantidad_excel, 1) <= tolerancia:
                        if clave not in duplicados_verificados:
                            resultados_totales.append([
                                r.FECHA, r.ProvNombre, r.FactEntPedimento, r.FactEntNofact,
                                r.FactEntFolio, r.Caja_CTR, mat_encontrado, r.MatDescr, cantidad_siadal,
                                mat, cantidad_excel
                            ])
                            duplicados_verificados.add(clave)
                            self.materiales_encontrados_avanzados.add(mat)

            return True, "Búsqueda avanzada completada.", resultados_totales

        except Exception as e:
            return False, f"Error: {e}", []

    def obtener_materiales_no_encontrados(self, tipo_operacion):
        """Devuelve lista de materiales no encontrados después de consultar SQL."""
        if self.df_original is None or not self.materiales_siadal:
            return []

        df_filtrado = self.filtrar_datos(tipo_operacion)
        no_encontrados = df_filtrado[
            ~df_filtrado["Número Material"].astype(str).str.strip().isin(self.materiales_siadal)
        ]
        return no_encontrados["Número Material"].astype(str).str.strip().unique().tolist()

    def guardar_datos(self, datos, archivo):
        try:
            columnas = [
                "Fecha", "Proveedor", "Pedimento", "Factura",
                "Folio", "Caja", "Número Material", "Descripción", "Cantidad BD",
                "Material Entrada", "Cantidad Entrada"
            ]
            df_guardar = pd.DataFrame(datos, columns=columnas)
            df_guardar.to_excel(archivo, index=False)
            return True, "Datos guardados correctamente."
        except Exception as e:
            return False, f"No se pudo guardar el archivo: {e}"

    def reinyectar_coincidencias(self, resultados):
        nuevos = {str(f[6]).strip() for f in resultados if f[6]}
        if not nuevos:
            return False, "No hay coincidencias para reinyectar."

        self.materiales_siadal.update(nuevos)
        self.materiales_encontrados_avanzados.update(nuevos)
        return True, f"Se reinyectaron {len(nuevos)} materiales."

    def escribir_resultados_en_excel(self, ruta_entrada, ruta_salida, resultados_avanzados):
        try:
            wb = load_workbook(ruta_entrada)
            hoja = wb["DETALLE FAC"]

            # Buscar columnas necesarias
            encabezados = [cell.value for cell in hoja[1]]
            if "Número Material" not in encabezados or "Número Factura" not in encabezados:
                return False, "No se encontró alguna de las columnas necesarias en el archivo."

            col_material = encabezados.index("Número Material") + 1
            col_factura = encabezados.index("Número Factura") + 1
            nueva_col = col_material + 1

            # Escribir encabezado para columna nueva
            hoja.cell(row=1, column=nueva_col, value="Número Material SIADAL")

            # Construir diccionario clave=(material, factura)
            coincidencias_dict = {
                (str(r[9]).strip(), str(r[3]).strip()): str(r[6]).strip()
                for r in resultados_avanzados
            }

            # Limpieza del fondo verde si se encuentra coincidencia
            for fila in range(2, hoja.max_row + 1):
                num_mat = hoja.cell(row=fila, column=col_material).value
                factura = hoja.cell(row=fila, column=col_factura).value

                if num_mat and factura:
                    clave = (str(num_mat).strip(), str(factura).strip())
                    valor_siadal = coincidencias_dict.get(clave, "")

                    hoja.cell(row=fila, column=nueva_col, value=valor_siadal)

                    # Si se encontró coincidencia, eliminar color verde
                    if valor_siadal:
                        hoja.cell(row=fila, column=col_material).fill = PatternFill(fill_type=None)

            wb.save(ruta_salida)
            return True, "Archivo actualizado correctamente."

        except Exception as e:
            return False, f"Error al escribir en Excel: {e}"